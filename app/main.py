from __future__ import annotations

import io
import os
import re as _re
import shutil
import time as _time
import uuid
from io import BytesIO
from datetime import date as _date, datetime as _datetime
from pathlib import Path

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from app.firestore_service import FirestoreQuestService


class QuestCreate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(min_length=1, max_length=240)
    points: int = Field(ge=1, le=999)
    difficulty: str = Field(default="easy", pattern="^(easy|medium|hard)$")
    category: str = Field(default="other", pattern="^(daily|weekly|event|other)$")
    published_date: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}-\d{2}$")
    due_date: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}-\d{2}$")


class QuestUpdate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(min_length=1, max_length=240)
    points: int = Field(ge=1, le=999)
    difficulty: str = Field(default="easy", pattern="^(easy|medium|hard)$")
    category: str = Field(default="other", pattern="^(daily|weekly|event|other)$")
    published_date: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}-\d{2}$")
    due_date: str = Field(default="", pattern=r"^$|^\d{4}-\d{2}-\d{2}$")


class RewardCreate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(min_length=1, max_length=240)
    cost_points: int = Field(ge=1, le=9999)
    image_path: str = Field(default="")


class RewardUpdate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(min_length=1, max_length=240)
    cost_points: int = Field(ge=1, le=9999)
    image_path: str = Field(default="")


class ClaimStatusUpdate(BaseModel):
    status: str = Field(pattern="^(claimed|delivered|completed)$")


class CategoryBonusClaim(BaseModel):
    quest_ids: list[str] = Field(min_length=1)
    bonus_points: int = Field(ge=1, le=9999)


class AnnouncementCreate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    content: str = Field(min_length=1, max_length=400)
    ann_type: str = Field(default="system", pattern="^(system|event)$")
    event_time: str = Field(default="", max_length=80)


class QuestProofSubmit(BaseModel):
    image_path: str = Field(min_length=1)


class QuestReviewUpdate(BaseModel):
    approved: bool
    review_note: str = Field(default="", max_length=240)


class GiftboxSendCreate(BaseModel):
    user_id: str = Field(min_length=1, max_length=80)
    title: str = Field(min_length=1, max_length=80)
    content: str = Field(min_length=1, max_length=400)
    points: int = Field(default=0, ge=0, le=99999)


class GiftboxDeleteReadRequest(BaseModel):
    force: bool = False


class QuestBatchDeleteRequest(BaseModel):
    quest_ids: list[str] = Field(min_length=1)


class QuestTemplateCreate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(min_length=1, max_length=240)
    points: int = Field(ge=1, le=999)
    difficulty: str = Field(default="easy", pattern="^(easy|medium|hard)$")
    category: str = Field(default="other", pattern="^(daily|weekly|event|other)$")
    due_days: int = Field(default=7, ge=0, le=365)


class QuestRefreshSettingsUpdate(BaseModel):
    daily_count: int = Field(ge=0, le=30)
    weekly_count: int = Field(ge=0, le=30)


class DailyJournalUpsert(BaseModel):
    log_date: str = Field(pattern=r"^\d{4}-\d{2}-\d{2}$")
    visited_place: str = Field(default="", max_length=200)
    note: str = Field(default="", max_length=400)
    completed_quest_ids: list[str] = Field(default_factory=list)


class EventScheduleCreate(BaseModel):
    title: str = Field(min_length=1, max_length=80)
    description: str = Field(default="", max_length=400)
    start_date: str = Field(pattern=r"^\d{4}-\d{2}-\d{2}$")
    end_date: str = Field(pattern=r"^\d{4}-\d{2}-\d{2}$")
    point_multiplier: float = Field(default=1.0, ge=0.5, le=10.0)
    announcement_title: str = Field(default="", max_length=80)
    announcement_content: str = Field(default="", max_length=400)
    announcement_event_time: str = Field(default="", max_length=80)
    auto_announce: bool = Field(default=True)


class WishCreate(BaseModel):
    url: str = Field(min_length=1, max_length=1000)
    item_name: str = Field(default="", max_length=120)
    note: str = Field(default="", max_length=300)


class GachaDrawRequest(BaseModel):
    chest_type: str = Field(pattern="^(bronze|silver|gold)$")


load_dotenv()
project_root = Path(__file__).resolve().parents[1]

credential_dict: dict | None = None
credential_path: str | None = None

# Priority 1: Render Secret File path
# - Use FIREBASE_SERVICE_ACCOUNT_FILE for explicit file path
# - Fallback to common Render secret location
secret_file_path = os.getenv("FIREBASE_SERVICE_ACCOUNT_FILE", "/etc/secrets/firebase-service-account.json")
secret_candidate = Path(secret_file_path)
if secret_candidate.exists():
    credential_path = str(secret_candidate)
else:
    # Priority 2: Render env var JSON
    _cred_json_str = os.getenv("FIREBASE_CREDENTIAL_JSON")
    if _cred_json_str:
        import json as _json
        credential_dict = _json.loads(_cred_json_str)
    else:
        # Priority 3: local dev file path
        credential_setting = os.getenv("FIREBASE_SERVICE_ACCOUNT") or os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
        if credential_setting:
            candidate = Path(credential_setting)
            if not candidate.is_absolute():
                candidate = project_root / candidate
            credential_path = str(candidate)

service = FirestoreQuestService(
    project_id=os.getenv("GOOGLE_CLOUD_PROJECT") or None,
    credential_path=credential_path,
    credential_dict=credential_dict,
)

app = FastAPI(title="Love Quest", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=project_root / "static"), name="static")
app.mount("/img", StaticFiles(directory=project_root / "img"), name="img")


@app.get("/")
def index() -> FileResponse:
    return FileResponse(project_root / "static" / "index.html")


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/quests")
def list_quests(include_future: bool = Query(default=False)) -> list[dict]:
    return service.get_quests(include_future=include_future)


@app.post("/api/quests")
def create_quest(payload: QuestCreate) -> dict:
    return service.create_quest(payload.model_dump())


@app.put("/api/quests/{quest_id}")
def update_quest(quest_id: str, payload: QuestUpdate) -> dict:
    result = service.update_quest(quest_id=quest_id, payload=payload.model_dump())
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "任務更新失敗")})
    return result


@app.delete("/api/quests/{quest_id}")
def delete_quest(quest_id: str) -> dict:
    result = service.deactivate_quest(quest_id=quest_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "任務刪除失敗")})
    return result


@app.post("/api/quests/{quest_id}/promote")
def promote_quest(quest_id: str) -> dict:
    """Move quest to top of visible pool (sets high sort_order)."""
    result = service.set_quest_sort_order(quest_id, int(_time.time()))
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "操作失敗")})
    return result


@app.post("/api/quests/{quest_id}/demote")
def demote_quest(quest_id: str) -> dict:
    """Remove quest from visible pool (sets sort_order = -1 so it sorts last)."""
    result = service.set_quest_sort_order(quest_id, -1)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "操作失敗")})
    return result
    return result


@app.post("/api/quests/batch-delete")
def batch_delete_quests(payload: QuestBatchDeleteRequest) -> dict:
    return service.batch_deactivate_quests(payload.quest_ids)


@app.post("/api/quests/delete-all")
def delete_all_quests() -> dict:
    return service.deactivate_all_quests()


@app.get("/api/quests/deleted-recent")
def list_recent_deleted_quests() -> list[dict]:
    return service.get_recent_deleted_quests()


@app.post("/api/quests/deleted-recent/{log_id}/restore")
def restore_deleted_quest(log_id: str) -> dict:
    result = service.restore_deleted_quest(log_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail={"message": result.get("message")})
    return result


@app.delete("/api/quests/deleted-recent/all")
def permanently_delete_all_deleted_quests() -> dict:
    return service.permanently_delete_all_deleted_quests()


@app.delete("/api/quests/deleted-recent/{log_id}")
def permanently_delete_deleted_quest(log_id: str) -> dict:
    result = service.permanently_delete_deleted_quest(log_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail={"message": result.get("message")})
    return result


@app.get("/api/quest-templates")
def list_quest_templates() -> list[dict]:
    return service.get_quest_templates()


@app.post("/api/quest-templates")
def create_quest_template(payload: QuestTemplateCreate) -> dict:
    return service.create_quest_template(payload.model_dump())


@app.delete("/api/quest-templates/{template_id}")
def delete_quest_template(template_id: str) -> dict:
    return service.deactivate_quest_template(template_id)


@app.post("/api/quest-templates/{template_id}/spawn")
def spawn_quest_from_template(template_id: str) -> dict:
    result = service.spawn_quest_from_template(template_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "任務建立失敗")})
    return result


@app.get("/api/quest-refresh-settings")
def get_quest_refresh_settings() -> dict:
    return service.get_quest_refresh_settings()


@app.post("/api/quest-refresh-settings")
def update_quest_refresh_settings(payload: QuestRefreshSettingsUpdate) -> dict:
    return service.update_quest_refresh_settings(payload.model_dump())


@app.post("/api/quests/refresh-from-db")
def refresh_quests_from_db() -> dict:
    return service.refresh_quests_from_templates()


@app.get("/api/rewards")
def list_rewards() -> list[dict]:
    return service.get_rewards()


@app.post("/api/rewards")
def create_reward(payload: RewardCreate) -> dict:
    return service.create_reward(payload.model_dump())


@app.put("/api/rewards/{reward_id}")
def update_reward(reward_id: str, payload: RewardUpdate) -> dict:
    result = service.update_reward(reward_id=reward_id, payload=payload.model_dump())
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "獎勵更新失敗")})
    return result


@app.delete("/api/rewards/{reward_id}")
def delete_reward(reward_id: str) -> dict:
    result = service.deactivate_reward(reward_id=reward_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "獎勵刪除失敗")})
    return result


@app.post("/api/uploads/reward-image")
def upload_reward_image(file: UploadFile = File(...)) -> dict[str, str]:
    if not file.filename:
        raise HTTPException(status_code=400, detail={"message": "請選擇圖片"})

    content_type = file.content_type or ""
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail={"message": "只接受圖片檔案"})

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail={"message": "圖片格式僅支援 png/jpg/jpeg/webp/gif"})

    upload_dir = project_root / "img" / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_name = f"reward_{uuid.uuid4().hex}{suffix}"
    saved_path = upload_dir / saved_name
    with saved_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    return {"image_path": f"/img/uploads/{saved_name}"}


@app.post("/api/uploads/proof-image")
def upload_proof_image(file: UploadFile = File(...)) -> dict[str, str]:
    if not file.filename:
        raise HTTPException(status_code=400, detail={"message": "請選擇圖片"})

    content_type = file.content_type or ""
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail={"message": "只接受圖片檔案"})

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail={"message": "圖片格式僅支援 png/jpg/jpeg/webp/gif"})

    upload_dir = project_root / "img" / "proofs"
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_name = f"proof_{uuid.uuid4().hex}{suffix}"
    saved_path = upload_dir / saved_name
    with saved_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    return {"image_path": f"/img/proofs/{saved_name}"}


@app.get("/api/progress/{user_id}")
def get_progress(user_id: str) -> dict:
    return service.get_progress(user_id)


@app.get("/api/gacha/pity/{user_id}")
def get_gacha_pity(user_id: str) -> dict:
    return service.get_gacha_pity(user_id)


@app.post("/api/gacha/draw/{user_id}")
def draw_gacha(user_id: str, payload: GachaDrawRequest) -> dict:
    result = service.draw_gacha(user_id=user_id, chest_type=payload.chest_type)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "抽獎失敗")})
    return result


@app.post("/api/quests/{quest_id}/accept/{user_id}")
def accept_quest(quest_id: str, user_id: str) -> dict:
    result = service.accept_quest(user_id=user_id, quest_id=quest_id)
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message})
    return {"message": result.message}


@app.post("/api/quests/{quest_id}/abandon/{user_id}")
def abandon_quest(quest_id: str, user_id: str) -> dict:
    result = service.abandon_quest(user_id=user_id, quest_id=quest_id)
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message})
    return {"message": result.message}


@app.post("/api/quests/{quest_id}/submit/{user_id}")
def submit_quest(quest_id: str, user_id: str, payload: QuestProofSubmit) -> dict:
    result = service.submit_quest_proof(
        user_id=user_id,
        quest_id=quest_id,
        image_path=payload.image_path,
    )
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message})
    return {"message": result.message}


@app.get("/api/quest-submissions")
def list_quest_submissions() -> list[dict]:
    return service.get_quest_submissions()


@app.post("/api/quest-submissions/{user_id}/{quest_id}/review")
def review_quest_submission(user_id: str, quest_id: str, payload: QuestReviewUpdate) -> dict:
    result = service.review_quest_submission(
        user_id=user_id,
        quest_id=quest_id,
        approved=payload.approved,
        review_note=payload.review_note,
    )
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "審核失敗")})
    return result


@app.post("/api/quests/{quest_id}/complete/{user_id}")
def complete_quest(quest_id: str, user_id: str) -> dict:
    result = service.complete_quest(user_id=user_id, quest_id=quest_id)
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message, "points": result.points})
    return {"message": result.message, "points": result.points}


@app.post("/api/rewards/{reward_id}/claim/{user_id}")
def claim_reward(reward_id: str, user_id: str) -> dict:
    result = service.claim_reward(user_id=user_id, reward_id=reward_id)
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message, "points": result.points})
    return {"message": result.message, "points": result.points}


@app.get("/api/claims")
def list_claims() -> list[dict]:
    return service.get_claims()


@app.get("/api/claims/{user_id}")
def list_claims_by_user(user_id: str) -> list[dict]:
    return service.get_claims_by_user(user_id)


@app.post("/api/claims/{user_id}/{reward_id}/status")
def update_claim_status(user_id: str, reward_id: str, payload: ClaimStatusUpdate) -> dict:
    result = service.update_claim_status(user_id=user_id, reward_id=reward_id, status=payload.status)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "更新失敗")})
    return result


@app.post("/api/category-bonus/{category}/{user_id}")
def claim_category_bonus(category: str, user_id: str, payload: CategoryBonusClaim) -> dict:
    result = service.claim_category_bonus(
        user_id=user_id,
        category=category,
        quest_ids=payload.quest_ids,
        bonus_points=payload.bonus_points,
    )
    if not result.ok:
        raise HTTPException(status_code=400, detail={"message": result.message})
    return {"message": result.message, "points": result.points}


@app.get("/api/announcements")
def list_announcements() -> list[dict]:
    return service.get_announcements()


@app.post("/api/announcements")
def create_announcement(payload: AnnouncementCreate) -> dict:
    return service.create_announcement(payload.model_dump())


@app.delete("/api/announcements/{announcement_id}")
def delete_announcement(announcement_id: str) -> dict:
    result = service.delete_announcement(announcement_id=announcement_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "刪除失敗")})
    return result


@app.post("/api/giftbox/send")
def send_giftbox_mail(payload: GiftboxSendCreate) -> dict:
    return service.create_giftbox_mail(payload.model_dump())


@app.get("/api/giftbox/{user_id}")
def list_giftbox_mails(user_id: str) -> list[dict]:
    return service.get_giftbox_mails(user_id)


@app.post("/api/giftbox/{user_id}/{mail_id}/claim")
def claim_giftbox_attachment(user_id: str, mail_id: str) -> dict:
    result = service.claim_giftbox_attachment(user_id=user_id, mail_id=mail_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "領取失敗")})
    return result


@app.post("/api/giftbox/{user_id}/claim-all")
def claim_all_giftbox_attachments(user_id: str) -> dict:
    return service.claim_all_giftbox_attachments(user_id)


@app.post("/api/giftbox/{user_id}/mark-all-read")
def mark_all_giftbox_read(user_id: str) -> dict:
    return service.mark_all_giftbox_read(user_id)


@app.post("/api/giftbox/{user_id}/delete-read")
def delete_read_giftbox_mails(user_id: str, payload: GiftboxDeleteReadRequest) -> dict:
    return service.delete_read_giftbox_mails(user_id=user_id, force=payload.force)


@app.get("/api/giftbox-history")
def list_giftbox_history() -> list[dict]:
    return service.get_giftbox_send_history()


@app.get("/api/completed-quests/{user_id}")
def list_completed_quests_by_date(user_id: str, date: str) -> list[dict]:
    return service.get_completed_quests_by_date(user_id=user_id, log_date=date)


@app.get("/api/daily-journal/{user_id}")
def get_daily_journal(user_id: str, date: str) -> dict:
    return service.get_daily_journal(user_id=user_id, log_date=date)


@app.post("/api/daily-journal/{user_id}")
def upsert_daily_journal(user_id: str, payload: DailyJournalUpsert) -> dict:
    return service.upsert_daily_journal(user_id=user_id, payload=payload.model_dump())


# ── Event Schedules ──────────────────────────────────────────

@app.get("/api/event-schedules")
def list_event_schedules() -> list[dict]:
    return service.get_event_schedules()


@app.post("/api/event-schedules")
def create_event_schedule(payload: EventScheduleCreate) -> dict:
    return service.create_event_schedule(payload.model_dump())


@app.delete("/api/event-schedules/{event_id}")
def delete_event_schedule(event_id: str) -> dict:
    result = service.delete_event_schedule(event_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "刪除失敗")})
    return result


# ── Wish Pool ─────────────────────────────────────────────────

@app.get("/api/wishes")
def list_wishes(user_id: str | None = None) -> list[dict]:
    return service.get_wishes(user_id=user_id)


@app.post("/api/wishes")
def create_wish(payload: WishCreate, user_id: str = "郭芸甄") -> dict:
    return service.create_wish(user_id=user_id, payload=payload.model_dump())


@app.patch("/api/wishes/{wish_id}/fulfill")
def fulfill_wish(wish_id: str) -> dict:
    result = service.fulfill_wish(wish_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "操作失敗")})
    return result


@app.delete("/api/wishes/{wish_id}")
def delete_wish(wish_id: str) -> dict:
    result = service.delete_wish(wish_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail={"message": result.get("message", "刪除失敗")})
    return result


# ── XLSX Template Download ────────────────────────────────────

_DATE_RE = _re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _normalize_excel_date(value) -> str:
    """Normalize Excel cell date value to YYYY-MM-DD; return __INVALID__ for bad non-empty input."""
    if value is None:
        return ""
    if isinstance(value, _datetime):
        return value.date().isoformat()
    if isinstance(value, _date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    if _DATE_RE.match(text):
        return text
    # Handle Excel datetime stringified values like "2026-05-01 00:00:00"
    m = _re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)
    return "__INVALID__"


def _style_header_row(ws, bg_hex: str) -> None:
    fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = max(max_len + 4, 18)


@app.get("/api/template/quests-events.xlsx")
def download_quest_event_template() -> StreamingResponse:
    wb = Workbook()

    # Sheet 1 – batch quest import
    ws1 = wb.active
    ws1.title = "任務批次新增"
    ws1.append(["任務名稱*", "任務描述*", "點數*", "難度(easy/medium/hard)", "種類(daily/weekly/event/other)", "發布日期(YYYY-MM-DD)", "截止日期(YYYY-MM-DD)"])
    ws1.append(["每日：刷牙洗臉", "每天早晚各刷牙一次", 10, "easy", "daily", "", ""])
    ws1.append(["每週：整理房間", "每週至少整理一次房間", 25, "medium", "weekly", "", ""])
    ws1.append(["活動：試吃新品", "到超商試吃活動新品", 40, "medium", "event", "2026-05-01", "2026-05-31"])
    _style_header_row(ws1, "5E3023")

    # Sheet 2 – event schedule
    ws2 = wb.create_sheet("活動排程")
    ws2.append(["活動名稱*", "活動說明", "開始日期*(YYYY-MM-DD)", "結束日期*(YYYY-MM-DD)",
                "點數倍率(預設1.0)", "自動發布公告(TRUE/FALSE)", "公告標題", "公告內容"])
    ws2.append(["五月甜蜜活動", "五月份特別活動，完成任務點數加倍！",
                "2026-05-01", "2026-05-31", 2.0, "TRUE", "🎉 五月活動開跑！",
                "五月份完成任務可獲得雙倍點數，快來參加！"])
    _style_header_row(ws2, "1B5E20")

    # Sheet 3 – event-only quests
    ws3 = wb.create_sheet("活動專屬任務")
    ws3.append(["任務名稱*", "任務描述*", "點數*", "難度(easy/medium/hard)",
                "發布日期(YYYY-MM-DD)", "截止日期(YYYY-MM-DD)"])
    ws3.append(["拍照打卡", "到活動地點拍照上傳打卡", 50, "medium", "2026-05-01", "2026-05-31"])
    _style_header_row(ws3, "1A237E")

    # Sheet 4 – limited rewards
    ws4 = wb.create_sheet("限定商品")
    ws4.append(["商品名稱*", "商品描述*", "兌換點數*"])
    ws4.append(["限定版貼紙組", "五月限定可愛貼紙一組", 80])
    _style_header_row(ws4, "6A1B9A")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="quests_events_template.xlsx"'},
    )


@app.get("/api/template/event-quests.xlsx")
def download_event_quests_template() -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "活動專屬任務"
    ws.append([
        "任務名稱*",
        "任務描述*",
        "點數*",
        "難度(easy/medium/hard)",
        "發布日期(YYYY-MM-DD，可空白)",
        "截止日期(YYYY-MM-DD，可空白)",
    ])
    ws.append(["拍照打卡", "到活動地點拍照上傳打卡", 50, "medium", "2026-05-01", "2026-05-31"])
    ws.append(["活動問答", "完成活動知識問答", 20, "easy", "", ""])
    _style_header_row(ws, "1A237E")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="event_quests_template.xlsx"'},
    )


# ── XLSX Import – quests ──────────────────────────────────────

@app.post("/api/quests/import-xlsx")
def import_quests_xlsx(file: UploadFile = File(...)) -> dict:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail={"message": "請上傳 .xlsx 格式的檔案"})
    contents = file.file.read()
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    sheet_name = "任務批次新增"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=400, detail={"message": f'找不到工作表「{sheet_name}」，請使用官方範本'})
    ws = wb[sheet_name]
    created = 0
    errors: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        title = str(row[0]).strip()
        description = str(row[1] or "").strip()
        if not description:
            errors.append(f"第 {row_idx} 行：任務描述不可為空")
            continue
        try:
            points = max(1, min(999, int(row[2] or 10)))
        except (ValueError, TypeError):
            errors.append(f"第 {row_idx} 行：點數格式錯誤")
            continue
        difficulty = str(row[3] or "easy").strip().lower()
        if difficulty not in ("easy", "medium", "hard"):
            difficulty = "easy"
        category = str(row[4] or "other").strip().lower()
        if category not in ("daily", "weekly", "event", "other"):
            category = "other"
        pub_date = _normalize_excel_date(row[5] if len(row) > 5 else "")
        due_date = _normalize_excel_date(row[6] if len(row) > 6 else "")
        if pub_date == "__INVALID__":
            pub_date = ""
        if due_date == "__INVALID__":
            due_date = ""
        try:
            service.create_quest({"title": title, "description": description,
                                   "points": points, "difficulty": difficulty,
                                   "category": category, "published_date": pub_date,
                                   "due_date": due_date})
            created += 1
        except Exception as exc:
            errors.append(f"第 {row_idx} 行建立失敗：{exc}")
    return {
        "ok": True,
        "message": f"已匯入 {created} 筆任務" + (f"，{len(errors)} 筆錯誤" if errors else ""),
        "created": created,
        "errors": errors,
    }


# ── XLSX Import – event schedule ──────────────────────────────

@app.post("/api/event-schedules/import-xlsx")
def import_event_schedule_xlsx(file: UploadFile = File(...)) -> dict:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail={"message": "請上傳 .xlsx 格式的檔案"})
    contents = file.file.read()
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    quests_created = 0
    rows_seen = 0
    imported_quests: list[dict] = []
    errors: list[str] = []

    if "活動專屬任務" in wb.sheetnames:
        ws = wb["活動專屬任務"]
    else:
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None

    if ws is None:
        raise HTTPException(status_code=400, detail={"message": "xlsx 內沒有可讀取的工作表"})

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        title_raw = str(row[0] or "").strip()
        if not title_raw:
            continue
        # Skip accidental header rows copied into data region
        if "任務名稱" in title_raw:
            continue

        rows_seen += 1

        title = title_raw
        description = str(row[1] or "").strip()
        if not title:
            errors.append(f"第 {row_idx} 行：任務名稱不可為空")
            continue
        if not description:
            # Allow blank description; fallback to title for convenience
            description = title

        raw_points = row[2] if len(row) > 2 else 10
        try:
            points = max(1, min(999, int(raw_points or 10)))
        except (ValueError, TypeError):
            # Tolerate values like "20點" or "20.0"
            m = _re.search(r"-?\d+(?:\.\d+)?", str(raw_points or ""))
            if m:
                try:
                    points = max(1, min(999, int(float(m.group(0)))))
                except (ValueError, TypeError):
                    points = 10
            else:
                points = 10

        difficulty = str(row[3] or "easy").strip().lower()
        if difficulty not in ("easy", "medium", "hard"):
            difficulty = "easy"

        pub_date = _normalize_excel_date(row[4] if len(row) > 4 else "")
        due_date = _normalize_excel_date(row[5] if len(row) > 5 else "")
        if pub_date == "__INVALID__":
            errors.append(f"第 {row_idx} 行：發布日期格式無法辨識，已改為空白")
            pub_date = ""
        if due_date == "__INVALID__":
            errors.append(f"第 {row_idx} 行：截止日期格式無法辨識，已改為空白")
            due_date = ""

        prefixed_title = title if title.startswith("活動：") else f"活動：{title}"
        created = service.create_quest(
            {
                "title": prefixed_title,
                "description": description,
                "points": points,
                "difficulty": difficulty,
                "category": "event",
                "published_date": pub_date,
                "due_date": due_date,
            }
        )
        imported_quests.append(
            {
                "id": created.get("id"),
                "title": created.get("title", prefixed_title),
                "description": created.get("description", description),
                "points": int(created.get("points", points)),
                "difficulty": created.get("difficulty", difficulty),
                "published_date": created.get("published_date", pub_date),
                "due_date": created.get("due_date", due_date),
                "category": "event",
            }
        )
        quests_created += 1

    return {
        "ok": True,
        "message": f"已匯入活動任務 {quests_created} 筆" + (f"，{len(errors)} 筆錯誤" if errors else ""),
        "rows_seen": rows_seen,
        "quests_created": quests_created,
        "imported_quests": imported_quests,
        "errors": errors,
    }
